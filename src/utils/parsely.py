from openpyxl import load_workbook
import json
import os
from django.conf import settings

class Parsely:

    def __init__(self, fname, satname):
        self.filename = fname
        self.satname = satname
        self.json = {}


        print(self.filename, self.satname)
        print('I got here and I like parsely')

    def chop(self):

        pthlst = [settings.MEDIA_DIR, self.filename]
        fpth = os.path.join(*pthlst)
        print(fpth)
        confmem = load_workbook(filename = fpth, data_only=True)



        # AMC-1_AMC1303_2015.xlsx
        # AMC-2_2019_0048.xlsx
        # AMC-3_2018_0023.xlsx
        # AMC-4_2019_0023xls.xlsx
        # AMC-6_2019_0077.xlsx
        # AMC-7_2019_0022.xlsx
        # AMC-8_2019_0013.xlsx
        # AMC-11_2018_0013.xlsx
        # AMC-15_2017_0032.xlsx
        # AMC-16_AMC958_2013.xlsx
        # AMC-18_2019_0016.xlsx
        # AMC-21_2019_0022.xlsx
        # Ciel2_CMGxxx_2013.xlsm
        # NSS-6_2019_0064.xlsx
        # NSS-7_2018_0035.xlsx
        # NSS-9_2019_0013.xlsx
        # NSS-10_2018_0015.xlsx
        # NSS-11_2019_0043.xlsx
        # QS-1_Config_2014_04_11.xlsm
        # SES-1_2019_0024.xlsx
        # SES-2_2017_0020.xlsx
        # SES-3_2018_0014.xlsx
        # SES-8_2019_0035.xlsx
        # SES-9_2019_0077.xlsx
        # SES-11_2018_0021.xlsx
        # SES-15_2019_0023.xlsx

        print(confmem.sheetnames)
        #ws = confmem['New Config']
        #print(ws['D9'].value)


        sat = self.satname
        print('satname is: ' + sat)
        #sat = 'CIEL2'

        if sat == 'NSS6':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=5, max_col=19, max_row=5, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=9, max_col=19, max_row=100, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=9, max_col=19, max_row=100, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('NSS6')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


        #    for changes (build for loop to extract data from tuples contained in chg and create a string in standard format that can be exported as json
            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][16]
                a4[x] = chg[x][1]
                a5[x] = None
                a6[x] = chg[x][3]
                a7[x] = None
                a8[x] = chg[x][10]
                a9[x] = chg[x][8]
                a10[x] = chg[x][7]
                a11[x] = chg[x][6]
                a12[x] = None
                a13[x] = None
                a14[x] = chg[x][15]
                a15[x] = None
                if chg[x][14] == 'ON':
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'NSS7':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=7, max_col=20, max_row=7, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=11, max_col=20, max_row=125, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=11, max_col=20, max_row=125, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('NSS7')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][17]
                a4[x] = chg[x][1]
                if chg[x][12] == 'On':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][10]
                a9[x] = chg[x][8]
                a10[x] = chg[x][11]
                a11[x] = chg[x][7]
                a12[x] = None
                a13[x] = None
                a14[x] = chg[x][16]
                a15[x] = None
                if chg[x][15] == 'On':
                    a16[x] = "ALC"
                elif chg[x][15] == 'Off':
                    a16[x] = "FGM"
                else:
                    a16[x] = chg[x][15]


            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'NSS9':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=14, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=8, max_col=14, max_row=63, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=8, max_col=14, max_row=63, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('NSS9')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][1]
                a3[x] = chg[x][1]
                a4[x] = chg[x][0]
                if chg[x][10] == 'On':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][9]
                a9[x] = chg[x][8]
                a10[x] = chg[x][7]
                a11[x] = chg[x][6]
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                a16[x] = "FGM only spacecraft"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AAP01':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=19, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=19, max_row=43, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=19, max_row=43, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AAP01')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                a5[x] = chg[x][14]
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][7]
                a9[x] = chg[x][11]
                a10[x] = chg[x][10]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][8]
                a14[x] = chg[x][16]
                a15[x] = None
                if chg[x][15] == 'Off':
                    a16[x] = "FGM"
                else:
                    a16[x] = "ALC"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC01':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=16, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=16, max_row=59, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=16, max_row=59, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC01')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][11] == 'ON':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][10]
                a10[x] = chg[x][9]
                a11[x] = chg[x][8]
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                if chg[x][13] == 'ON':
                    a16[x] = "Limiter Mode"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC02':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=16, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=11, max_col=16, max_row=60, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=11, max_col=16, max_row=60, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC02')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][13]
                a4[x] = chg[x][1]
                if chg[x][10] == 'On':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][8]
                a9[x] = chg[x][6]
                a10[x] = chg[x][9]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                if chg[x][12] == 'OFF':
                    a16[x] = "Limiter Mode"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC03':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=59, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=12, max_col=15, max_row=61, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC03')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][11] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                if chg[x][12] == 'ON':
                    a16[x] = "Limiter Mode"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC04':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=63, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=12, max_col=15, max_row=65, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC04')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][10] == 'Off':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                if chg[x][12] == 'ON':
                    a16[x] = "Limiter Mode"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC06':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=63, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=12, max_col=15, max_row=65, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC06')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][10] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                if chg[x][12] == "ON":
                    a16[x] = "Limiter Mode"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'AMC07':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=59, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=14, max_col=15, max_row=63, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC07')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][10] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                a16[x] = "FGM only spacecraft"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'AMC08':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new=[]
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=59, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=12, max_col=15, max_row=61, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC08')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][10] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                a16[x] = "FGM only spacecraft"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC11':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=33, values_only=True):
                i=0
                satvals_new.append(new)
                print(satvals_new[i-1])
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=14, max_col=15, max_row=37, values_only=True):
                i=0
                satvals_old.append(old)
                print(satvals_old[i-1])
                i=i+1
            print('AMC11')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][10] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                a16[x] = "FGM only spacecraft"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC12':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=18, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=18, max_row=96, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=18, max_row=96, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC12')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][12] == 'Off':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][3]
                a7[x] = None
                a8[x] = chg[x][7]
                a9[x] = chg[x][11]
                a10[x] = chg[x][10]
                a11[x] = chg[x][9]
                a12[x] = None
                a13[x] = chg[x][8]
                a14[x] = None
                a15[x] = None
                if chg[x][12] == 'ON':
                    a16[x] = "Muted"
                else:
                    a16[x] = "Unmuted"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'AMC15':
            #Pull from worksheets and do comparison code here
            ws = confmem['All Transponders NEW']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=20, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=20, max_row=70, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old configuration']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=20, max_row=70, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC15')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))

            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][13] == 'On':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][6]
                a8[x] = chg[x][9]
                a9[x] = chg[x][12]
                a10[x] = chg[x][10]
                a11[x] = None
                a12[x] = chg[x][11]
                a13[x] = None
                a14[x] = None
                a15[x] = None
                a16[x] = chg[x][14]

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }


            change_json = json.dumps(changefile)

            with open('AMC15.txt', 'w') as json_file:
                json.dump(change_json, json_file)

        elif sat == 'AMC16':
            #Pull from worksheets and do comparison code here
            ws = confmem['Active Transponders']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=7, max_col=20, max_row=7, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=11, max_col=20, max_row=43, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old configuration']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=11, max_col=20, max_row=43, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC16')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][13] == 'Off':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][6]
                a8[x] = chg[x][9]
                a9[x] = chg[x][12]
                a10[x] = chg[x][10]
                a11[x] = None
                a12[x] = chg[x][11]
                a13[x] = None
                a14[x] = chg[x][16]
                a15[x] = None
                a16[x] = chg[x][14]

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'AMC18':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=33, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=14, max_col=15, max_row=37, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC18')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][10] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][9]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = None
                a16[x] = "FGM only spacecraft"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'AMC21':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=17, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=17, max_row=33, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=14, max_col=17, max_row=37, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('AMC21')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][11] == 'On':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][10]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][9]
                a14[x] = None
                a15[x] = chg[x][14]
                a16[x] = chg[x][13]

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'SES01':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=15, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=15, max_row=72, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=11, max_col=15, max_row=73, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES01')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][11] == 'Off':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][6]
                a9[x] = chg[x][10]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][9]
                a14[x] = None
                a15[x] = None
                if chg[x][12] == 'n/a':
                    a16[x] = "FGM"
                elif chg[x][12] == '-':
                    a16[x] = "FGM"
                else:
                    a16[x] = "ALC"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'SES02':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=20, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=20, max_row=72, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=13, max_col=20, max_row=75, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES02')
            print(set(satvals_new).difference(set(satvals_old)))

            changes = len(set(satvals_new).difference(set(satvals_old)))
            print(changes)

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][0]
                a4[x] = chg[x][1]
                if chg[x][13] == 'Off':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][6]
                a8[x] = chg[x][9]
                a9[x] = chg[x][12]
                a10[x] = chg[x][10]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][11]
                a14[x] = None
                a15[x] = None
                if chg[x][16] == 'n/a':
                    a16[x] = "FGM"
                elif chg[x][16] == '-':
                    a16[x] = "FGM"
                else:
                    a16[x] = "ALC"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'SES03':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=5, max_col=18, max_row=5, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=8, max_col=18, max_row=70, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=18, max_row=72, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES03')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][15]
                a4[x] = chg[x][1]
                if chg[x][11] == 'On':
                    a5[x] = "ALC"
                else:
                    a5[x] = "FGM"
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][7]
                a9[x] = chg[x][10]
                a10[x] = chg[x][8]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][9]
                a14[x] = None
                a15[x] = None
                if chg[x][14] == 'On':
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'SES08':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=6, max_col=19, max_row=6, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new =[]
            for new in ws.iter_rows(min_col=2, min_row=11, max_col=19, max_row=78, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=11, max_col=19, max_row=78, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES08')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][16]
                a4[x] = chg[x][1]
                if chg[x][12] == "On":
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][4]
                a7[x] = chg[x][5]
                a8[x] = chg[x][7]
                a9[x] = None
                a10[x] = chg[x][9]
                a11[x] = chg[x][11]
                a12[x] = None
                a13[x] = None
                a14[x] = chg[x][15]
                a15[x] = None
                if chg[x][13] == "On":
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'SES09':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=5, max_col=21, max_row=5, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=9, max_col=21, max_row=166, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=9, max_col=21, max_row=166, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES09')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][18]
                a4[x] = chg[x][1]
                if chg[x][14] == 'On':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][7]
                a9[x] = None
                a10[x] = chg[x][9]
                a11[x] = chg[x][13]
                a12[x] = None
                a13[x] = None
                a14[x] = chg[x][17]
                a15[x] = None
                if chg[x][15] == 'On':
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)


        elif sat == 'SES11':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=5, max_col=21, max_row=5, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=9, max_col=21, max_row=58, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=9, max_col=21, max_row=58, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES11')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][18]
                a4[x] = chg[x][1]
                if chg[x][14] == 'OFF':
                    a5[x] = "Muted"
                else:
                    a5[x] = "Unmuted"
                a6[x] = chg[x][3]
                a7[x] = chg[x][5]
                a8[x] = chg[x][7]
                a9[x] = chg[x][11]
                a10[x] = chg[x][9]
                a11[x] = chg[x][13]
                a12[x] = None
                a13[x] = None
                a14[x] = None
                a15[x] = chg[x][17]
                if chg[x][15] == "ON":
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'SES-15':
            #Pull from worksheets and do comparison code here
            ws = confmem['New CFG']
            top = []
            for hdg in ws.iter_rows(min_col=1, min_row=1, max_col=17, max_row=1, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=9, max_col=17, max_row=58, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old CFG']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=9, max_col=17, max_row=58, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('SES-15')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][15]
                a4[x] = None
                if chg[x][14] == 'ON':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][2]
                a7[x] = chg[x][4]
                a8[x] = chg[x][8]
                a9[x] = None
                a10[x] = chg[x][9]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][10]
                a14[x] = None
                a15[x] = chg[x][12]
                if chg[x][13] == 'ON':
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"


            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'CIEL2':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=7, max_col=30, max_row=7, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=30, max_row=75, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=30, max_row=75, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('CIEL2')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = None
                a4[x] = chg[x][3]
                if chg[x][16] == 'OFF':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = None
                a7[x] = chg[x][7]
                a8[x] = chg[x][12]
                a9[x] = None
                a10[x] = chg[x][20]
                a11[x] = chg[x][21]
                a12[x] = None
                a13[x] = chg[x][19]
                a14[x] = None
                a15[x] = chg[x][24]
                a16[x] = chg[x][17]

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        elif sat == 'QS1':
            #Pull from worksheets and do comparison code here
            ws = confmem['New Config']
            top = []
            for hdg in ws.iter_rows(min_col=2, min_row=7, max_col=22, max_row=7, values_only=True):
                i=0
                top.append(hdg)
                print(top[i-1])
                i=i+1
            satvals_new = []
            for new in ws.iter_rows(min_col=2, min_row=10, max_col=22, max_row=41, values_only=True):
                i=0
                satvals_new.append(new)
                i=i+1
            ws = confmem['Old Config']
            satvals_old = []
            for old in ws.iter_rows(min_col=2, min_row=10, max_col=22, max_row=41, values_only=True):
                i=0
                satvals_old.append(old)
                i=i+1
            print('QS1')
            print(set(satvals_new).difference(set(satvals_old)))
            changes = len(set(satvals_new).difference(set(satvals_old)))

            chg = list(set(satvals_new).difference(set(satvals_old)))


            a1 = [0] * changes
            a2 = [0] * changes
            a3 = [0] * changes
            a4 = [0] * changes
            a5 = [0] * changes
            a6 = [0] * changes
            a7 = [0] * changes
            a8 = [0] * changes
            a9 = [0] * changes
            a10 = [0] * changes
            a11 = [0] * changes
            a12 = [0] * changes
            a13 = [0] * changes
            a14 = [0] * changes
            a15 = [0] * changes
            a16 = [0] * changes

            for x in range(changes):
                a1[x] = x+1
                a2[x] = chg[x][0]
                a3[x] = chg[x][19]
                a4[x] = chg[x][1]
                if chg[x][13] == 'OFF':
                    a5[x] = "Unmuted"
                else:
                    a5[x] = "Muted"
                a6[x] = chg[x][3]
                a7[x] = chg[x][6]
                a8[x] = chg[x][9]
                a9[x] = None
                a10[x] = chg[x][12]
                a11[x] = None
                a12[x] = None
                a13[x] = chg[x][11]
                a14[x] = chg[x][17]
                a15[x] = None
                if chg[x][16] == "ON":
                    a16[x] = "ALC"
                else:
                    a16[x] = "FGM"

            changefile = {
                "Change ID":a1,
                "U/L Channel":a2,
                "D/L Channel":a3,
                "Service Status":a4,
                "Mute Status":a5,
                "Transponder ID":a6,
                "Receiver ID":a7,
                "HPA ID":a8,
                "SFD":a9,
                "FCA":a10,
                "OPA":a11,
                "IFP":a12,
                "GAIN":a13,
                "IBO":a14,
                "OBO":a15,
                "OPMOD":a16,
                }

            print(changefile)

            change_json = json.dumps(changefile)
            print(change_json)

        else:
            print('Error: Satellite not found')
            change_json = {}
        self.json = change_json
